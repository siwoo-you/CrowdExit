"""
서울시 주요 121장소 목록 서비스

data/seoul_121_areas.xlsx 파일을 읽어서
CrowdExit 출발지 검색용 데이터로 제공한다.
"""

from pathlib import Path
from typing import List, Dict, Any

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent

AREA_FILE = (
    BASE_DIR
    / "data"
    / "seoul_121_areas.xlsx"
)


# ============================================================
# 전체 121장소 읽기
# ============================================================

def load_areas() -> List[Dict[str, Any]]:

    if not AREA_FILE.exists():
        raise FileNotFoundError(
            f"121장소 파일을 찾을 수 없습니다: {AREA_FILE}"
        )

    workbook = load_workbook(
        AREA_FILE,
        read_only=True,
        data_only=True,
    )

    sheet = workbook[
        "장소목록"
    ]

    areas = []

    # 첫 행은 헤더이므로 제외
    rows = sheet.iter_rows(
        min_row=2,
        values_only=True,
    )

    for row in rows:

        category = row[0]
        number = row[1]
        area_code = row[2]
        area_name = row[3]
        english_name = row[4]

        if not area_code or not area_name:
            continue

        areas.append({
            "category": category,
            "number": number,
            "area_code": area_code,
            "area_name": area_name,
            "english_name": english_name,
        })

    return areas


# ============================================================
# 검색
# ============================================================

def search_areas(
    keyword: str = ""
) -> List[Dict[str, Any]]:

    areas = load_areas()

    keyword = str(
        keyword or ""
    ).strip().lower()

    if not keyword:
        return areas

    results = []

    for area in areas:

        area_name = str(
            area.get(
                "area_name",
                ""
            )
        ).lower()

        english_name = str(
            area.get(
                "english_name",
                ""
            )
        ).lower()

        category = str(
            area.get(
                "category",
                ""
            )
        ).lower()

        if (
            keyword in area_name
            or keyword in english_name
            or keyword in category
        ):
            results.append(
                area
            )

    return results


# ============================================================
# 코드로 조회
# ============================================================

def get_area_by_code(
    area_code: str
) -> Dict[str, Any] | None:

    area_code = str(
        area_code or ""
    ).strip()

    for area in load_areas():

        if (
            area.get(
                "area_code"
            )
            == area_code
        ):
            return area

    return None


# ============================================================
# 이름으로 조회
# ============================================================

def get_area_by_name(
    area_name: str
) -> Dict[str, Any] | None:

    area_name = str(
        area_name or ""
    ).strip()

    for area in load_areas():

        if (
            area.get(
                "area_name"
            )
            == area_name
        ):
            return area

    return None


# ============================================================
# 테스트
# ============================================================

if __name__ == "__main__":

    areas = load_areas()

    print(
        "전체 장소 수:",
        len(areas)
    )

    print()

    print(
        "잠실 검색 결과:"
    )

    for area in search_areas(
        "잠실"
    ):
        print(
            area
        )